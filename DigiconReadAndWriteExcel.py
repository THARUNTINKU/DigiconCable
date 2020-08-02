import xlrd
from openpyxl import load_workbook


def generateDictForCustID():
    workbook = xlrd.open_workbook("I:/Digicon/Digicon/DigiconTotalList.xlsx")
    # Read data from Excel sheet named "Check"
    sheet = workbook.sheet_by_name("DigiconTotalList")
    dict2 = {}
    for curr_row in range(1, sheet.nrows):
        row_data = []
        for curr_col in range(1, 3):
            # Read the data in the current cell
            data = sheet.cell_value(curr_row, curr_col)
            # print("curr_row", curr_row, "curr_col", curr_col, "Data: ", data)
            row_data.append(data)
            # x = txt.split("(", 1)
        if row_data[0] != "NULL" and row_data[1] != "N/A":
            # print("row_data[0]: ", row_data[0], "row_data[1]: ", row_data[1])
            serial = row_data[0]
            custid = row_data[1].split("(", 1)
            dict2[serial[1:]] = custid[0]

    mydict = dict(dict2)
    return mydict


def setlist(totaldict):
    workbook = xlrd.open_workbook("I:/Digicon/Digicon/INDRA_SAT_VISION.xlsx")
    wb = load_workbook("I:/Digicon/Digicon/INDRA_SAT_VISION.xlsx")

    # Read data from Excel sheet
    sheet_list = ["Kamaraj Nagar", "Etteiyampatti"]
    for sheetname in sheet_list:
        ws = wb[sheetname]
        sheet = workbook.sheet_by_name(sheetname)
        for curr_row in range(1, sheet.nrows):
            for curr_col in range(4, 6):
                # Read the data in the current cell
                if curr_col == 4:
                    data = sheet.cell_value(curr_row, curr_col)
                    # print("curr_row", curr_row, "curr_col", curr_col, "Data: ", data)
                    wcell1 = ws.cell(curr_row + 1, 6)
                    if data!="":
                        # print(data)
                        if data in totaldict:
                            cusid = totaldict[data]
                            # print(cusid)
                            wcell1.value = cusid

        wb.save("I:/Digicon/Digicon/INDRA_SAT_VISION.xlsx")
        print(sheetname, "customer id records saved successfully" )


totalCustIDdict = generateDictForCustID()
#print(totaldict)
#print("Filter count: ", len(totaldict))
totalCustIDlist = []
for i in totalCustIDdict:
    totalCustIDlist.append(i)
#print(totallist)
setlist(totalCustIDdict)
