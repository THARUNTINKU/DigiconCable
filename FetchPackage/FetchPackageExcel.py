import xlrd
from openpyxl import load_workbook


def generateActiveDict():
    workbook = xlrd.open_workbook("D:/Personal/Cable/Digicon/PackageWiseList.xlsx")
    # Read data from Excel sheet named "Active"
    sheet_name = workbook.sheet_by_name("Active")
    active_packages = {}
    for curr_row in range(1, sheet_name.nrows):
        row_data = []
        for curr_col in range(2, 4):
            # Read the data in the current cell
            data = sheet_name.cell_value(curr_row, curr_col)
            # print("curr_row", curr_row, "curr_col", curr_col, "Data: ", data)
            row_data.append(data)
            # x = txt.split("(", 1)

        # print("row_data[0]: ", row_data[0], "row_data[1]: ", row_data[1])
        vc_number = row_data[0]
        package = row_data[1]
        active_packages[vc_number[1:]] = package[1:]

    result = dict(active_packages)
    return result


def set_packagedetailslist(packages):
    workbook = xlrd.open_workbook("D:/Personal/Cable/Digicon/INDRA_SAT_VISION_Testing.xlsx")
    wb = load_workbook(r'D:/Personal/Cable/Digicon/INDRA_SAT_VISION_Testing.xlsx')

    # Read data from Excel sheet
    sheet_list = ["Vallur-1", "Vallur-2", "SankarK", "Pudhur", "Muslim St", "A.Colony", "Kamaraj Nagar",
                  "Etteiyampatti"]

    TotalBox = 0
    for sheet_name in sheet_list:
        ws = wb[sheet_name]
        sheet = workbook.sheet_by_name(sheet_name)
        for curr_row in range(1, sheet.nrows):
            for curr_col in range(3, 5):
                # Read the data in the current cell
                if curr_col == 4:
                    data = sheet.cell_value(curr_row, curr_col)
                    # print("curr_row", curr_row, "curr_col", curr_col, "Data: ", data)
                    packagecell = ws.cell(curr_row + 1, 4)
                    if data != '':
                        # print(data)
                        if data in packages:
                            packdet = packages[data]
                            # print(packdet)
                            TotalBox += 1
                            packagecell.value = packdet

        wb.save("D:/Personal/Cable/Digicon/INDRA_SAT_VISION_Testing.xlsx")
        print(sheet_name, " records saved successfully")

    print('TotalBox: ', TotalBox)